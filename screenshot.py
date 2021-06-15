import pyautogui as pg
import os
import statistics
from statistics import mode
import pyperclip
from PIL import Image, ImageEnhance, ImageFilter
import re
import cv2
import pytesseract
from pytesseract import Output
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl import load_workbook

pytesseract.pytesseract.tesseract_cmd = 'C:\\Program Files\\Tesseract-OCR\\tesseract.exe' 

# Create header
XLSX_NAME = "Player info.xlsx"
WORKSHEET_NAME = "Player Info"

# Save names
names_dict = {}

def setup_xlsx():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_NAME

    now = datetime.now()
    current_date_time = now.strftime("%d/%m/%Y %H:%M")

    worksheet["A1"] = "Created On:"
    worksheet["B1"] = f"{current_date_time}"

    worksheet["A2"] = "Rank"
    worksheet["B2"] = "Name"
    worksheet["C2"] = "UID"
    worksheet["D2"] = "Alliance"
    worksheet["E2"] = "Power"
    worksheet["F2"] = "T4 Kill Points"
    worksheet["G2"] = "T5 Kill Points"
    worksheet["H2"] = "Dead Troops"

    workbook.save(filename = XLSX_NAME)

    # Create player images folder if not exists
    if not os.path.exists("player images"):
        os.makedirs("player images")

############ Screenshot ############
def save_player_name(rank):
    wb = load_workbook(filename = XLSX_NAME)
    ws = wb[WORKSHEET_NAME]
    row_num = rank+2

    ws[f"A{row_num}"] = rank
    ws[f"B{row_num}"] = pyperclip.paste()

    wb.save(XLSX_NAME)

def capture_screenshot(y_position, count):
    os.chdir("player images")
    
    pg.click(x=1000, y=y_position, clicks=1, interval=1) # click the player

    pg.click(x=1343, y=422, clicks=1, interval=1) # move to ?(kill Points) and click
    pg.screenshot(f"{count}-kills.png")

    pg.click(x=470, y=800, clicks=1, interval=1) # move to more info and click
    pg.screenshot(f"{count}-death.png")

    os.chdir("..") # Move back to root folder

    pg.click(x=450, y=185, clicks=1, interval=1) # copy player name to clipboard
    save_player_name(count) # save player name to dictionary

    pg.click(x=1673, y=65, clicks=1, interval=1) # move to X and click
  
    pg.click(x=1637, y=123, clicks=1, interval=0.5) # move to X and click

    pg.moveTo(848, y_position, 0.5) # move to the id


def takeScreenshot(player_count: int):
    # Capture top 3
    count = 1
    top1_y = 340

    pg.click(x=1000, y=180, clicks=1, interval=0.5)

    while count <= 3:
        capture_screenshot(top1_y, count)
        count += 1
        top1_y += 120

    # Capture the rest
    UP_TO = player_count
    count = 4
    y_position = 700 # y position of 4th player
    while count <= UP_TO:
        capture_screenshot(y_position, count)

        y_position = 730
        count += 1


############ OCR ############

def enhance_Image(image):
    # Grayscale, Gaussian blur, Otsu's threshold
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (3,3), 0)
    thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]

    # Morph open to remove noise and invert image
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    opening = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    return opening

def enhance_Image2(image):
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(image,(5,5),0)
    image = cv2.threshold(blur,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)[1]
    kernel = np.ones((2,2),np.uint8)
    image = cv2.erode(image,kernel,iterations = 1)
    return image    

def remove_noise_and_smooth(image):
    image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    filtered = cv2.adaptiveThreshold(image.astype(np.uint8), 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 9, 41)
    kernel = np.ones((111, 111), np.uint8)
    opening = cv2.morphologyEx(filtered, cv2.MORPH_OPEN, kernel)
    closing = cv2.morphologyEx(opening, cv2.MORPH_CLOSE, kernel)
    or_image = cv2.bitwise_or(image, closing)
    return or_image

def write_to_xlsx(rank, data):
    wb = load_workbook(filename = XLSX_NAME)
    ws = wb[WORKSHEET_NAME]
    row_num = rank+2

    ws[f"C{row_num}"] = data["id"]
    ws[f"D{row_num}"] = data["alliance"]
    ws[f"E{row_num}"] = data["power"]
    ws[f"F{row_num}"] = data["t4_kills"]
    ws[f"G{row_num}"] = data["t5_kills"]
    ws[f"H{row_num}"] = data["dead"]

    wb.save(XLSX_NAME)

def ocr(player_count: int):
    PLAYER_COUNT = player_count
    counter = 1
    # data = [] # array of dictionaries (data_dict)
    while counter <= PLAYER_COUNT:
        data_dict = {}

        kill_image = cv2.imread(f"player images/{counter}-kills.png")

        # Player UID
        uid_img = kill_image[280:320, 935:1095]
        uid_img1 = enhance_Image(uid_img)
        uid_img2 = enhance_Image2(uid_img)
        uid_img3 = remove_noise_and_smooth(uid_img)

        # Perform OCR
        uid_list = []

        data = pytesseract.image_to_data(uid_img, output_type=Output.DICT)
        uid = "".join(data["text"])
        if uid.isdigit():
            uid_list.append(int(uid))
        else:
            uid = uid[0:-1] # remove ')'
            if uid.isdigit():
                uid_list.append(int(uid))

        data = pytesseract.image_to_data(uid_img1, output_type=Output.DICT)
        uid = "".join(data["text"])
        if uid.isdigit():
            uid_list.append(int(uid))
        else:
            uid = uid[0:-1] # remove ')'
            if uid.isdigit():
                uid_list.append(int(uid))

        data = pytesseract.image_to_data(uid_img2, output_type=Output.DICT)
        uid = "".join(data["text"])
        if uid.isdigit():
            uid_list.append(int(uid))
        else:
            uid = uid[0:-1] # remove ')'
            if uid.isdigit():
                uid_list.append(int(uid))

        data = pytesseract.image_to_data(uid_img3, output_type=Output.DICT)
        uid = "".join(data["text"])
        if uid.isdigit():
            uid_list.append(int(uid))
        else:
            uid = uid[0:-1] # remove ')'
            if uid.isdigit():
                uid_list.append(int(uid))
        
        # Get most frequent element
        if len(uid_list) > 0:
            player_id = mode(uid_list)
        else:
            player_id = uid

        print(f"id: {player_id}")
        data_dict["id"] = player_id

        # Alliance
        alliance_img = kill_image[434:475 , 770:921]
        alliance_img1 = enhance_Image(alliance_img)
        alliance_img2 = enhance_Image2(alliance_img)
        alliance_img3 = remove_noise_and_smooth(alliance_img)

        # Perform OCR
        alliance_list = []

        data = pytesseract.image_to_data(alliance_img, output_type=Output.DICT)
        alliance = "".join(data["text"])
        end_index = alliance.find(']')
        if end_index != -1 and len(alliance) > end_index:
            alliance = alliance[:end_index+1]
            alliance_list.append(alliance)

        data = pytesseract.image_to_data(alliance_img1, output_type=Output.DICT)
        alliance = "".join(data["text"])
        end_index = alliance.find(']')
        if end_index != -1 and len(alliance) > end_index:
            alliance = alliance[:end_index+1]
            alliance_list.append(alliance)

        data = pytesseract.image_to_data(alliance_img2, output_type=Output.DICT)
        alliance = "".join(data["text"])
        end_index = alliance.find(']')
        if end_index != -1 and len(alliance) > end_index:
            alliance = alliance[:end_index+1]
            alliance_list.append(alliance)

        data = pytesseract.image_to_data(alliance_img3, output_type=Output.DICT)
        alliance = "".join(data["text"])
        end_index = alliance.find(']')
        if end_index != -1 and len(alliance) > end_index:
            alliance = alliance[:end_index+1]
            alliance_list.append(alliance)
        
        if len(alliance_list) > 0:
            alliance = mode(alliance_list)

        print(f"alliance: {alliance}")
        data_dict["alliance"] = alliance

        # T4 Kills
        t4_img = kill_image[868:918, 1504:1713]
        t4_img1 = enhance_Image(t4_img)
        t4_img2 = enhance_Image2(t4_img)
        t4_img3 = remove_noise_and_smooth(t4_img)

        # Perform OCR
        t4_kills_list = []

        data = pytesseract.image_to_data(t4_img, output_type=Output.DICT)
        t4_kills = "".join(data["text"])
        t4_kills = t4_kills.replace(",", "")
        if t4_kills.isdigit():
            t4_kills_list.append(int(t4_kills))

        data = pytesseract.image_to_data(t4_img1, output_type=Output.DICT)
        t4_kills = "".join(data["text"])
        t4_kills = t4_kills.replace(",", "")
        if t4_kills.isdigit():
            t4_kills_list.append(int(t4_kills))

        data = pytesseract.image_to_data(t4_img2, output_type=Output.DICT)
        t4_kills = "".join(data["text"])
        t4_kills = t4_kills.replace(",", "")
        if t4_kills.isdigit():
            t4_kills_list.append(int(t4_kills))

        data = pytesseract.image_to_data(t4_img3, output_type=Output.DICT)
        t4_kills = "".join(data["text"])
        t4_kills = t4_kills.replace(",", "")
        if t4_kills.isdigit():
            t4_kills_list.append(int(t4_kills))
        
        if len(t4_kills_list) > 0:
            t4_kills = mode(t4_kills_list)
        else:
            t4_kills = -1

        print(f"t4 kills: {t4_kills}")
        data_dict["t4_kills"] = t4_kills

        # T5 Kills
        t5_img = kill_image[922:972, 1504:1713]
        t5_img1 = enhance_Image(t5_img)
        t5_img2 = enhance_Image2(t5_img)
        t5_img3 = remove_noise_and_smooth(t5_img)

        # Perform OCR
        t5_kills_list = []

        data = pytesseract.image_to_data(t5_img, output_type=Output.DICT)
        t5_kills = "".join(data["text"])
        t5_kills = t5_kills.replace(",", "")
        if t5_kills.isdigit():
            t5_kills_list.append(int(t5_kills))

        data = pytesseract.image_to_data(t5_img1, output_type=Output.DICT)
        t5_kills = "".join(data["text"])
        t5_kills = t5_kills.replace(",", "")
        if t5_kills.isdigit():
            t5_kills_list.append(int(t5_kills))

        data = pytesseract.image_to_data(t5_img2, output_type=Output.DICT)
        t5_kills = "".join(data["text"])
        t5_kills = t5_kills.replace(",", "")
        if t5_kills.isdigit():
            t5_kills_list.append(int(t5_kills))

        data = pytesseract.image_to_data(t5_img3, output_type=Output.DICT)
        t5_kills = "".join(data["text"])
        t5_kills = t5_kills.replace(",", "")
        if t5_kills.isdigit():
            t5_kills_list.append(int(t5_kills))
        
        if len(t5_kills_list) > 0:
            t5_kills = mode(t5_kills_list)
        else:
            t5_kills = -1
            
        print(f"t5 kills: {t5_kills}")
        data_dict["t5_kills"] = t5_kills

        
        ##########################################

        detail_image = cv2.imread(f"player images/{counter}-death.png")

        # Power
        power_img = detail_image[154:214 , 874:1162]
        power_img1 = enhance_Image(power_img)
        power_img2 = enhance_Image2(power_img)
        power_img3 = remove_noise_and_smooth(power_img)

        # Perform OCR
        power_list = []

        data = pytesseract.image_to_data(power_img, output_type=Output.DICT)
        power = "".join(data["text"])
        power = power.replace(",", "").replace("Power:", "")
        if power.isdigit():
            power_list.append(int(power))

        data = pytesseract.image_to_data(power_img1, output_type=Output.DICT)
        power = "".join(data["text"])
        power = power.replace(",", "").replace("Power:", "")
        if power.isdigit():
            power_list.append(int(power))

        data = pytesseract.image_to_data(power_img2, output_type=Output.DICT)
        power = "".join(data["text"])
        power = power.replace(",", "").replace("Power:", "")
        if power.isdigit():
            power_list.append(int(power))

        data = pytesseract.image_to_data(power_img3, output_type=Output.DICT)
        power = "".join(data["text"])
        power = power.replace(",", "").replace("Power:", "")
        if power.isdigit():
            power_list.append(int(power))

        if len(power_list) > 0:
            power = mode(power_list)

        print(f"power: {power}")
        data_dict["power"] = power

        # Dead
        dead_img = detail_image[533:593 ,1387:1587]
        dead_img1 = enhance_Image(dead_img)
        dead_img2 = enhance_Image2(dead_img)
        dead_img3 = remove_noise_and_smooth(dead_img)

        # Perform OCR
        dead_list = []

        data = pytesseract.image_to_data(dead_img, output_type=Output.DICT)
        dead = "".join(data["text"])
        dead = dead.replace(",", "")
        if dead.isdigit():
            dead_list.append(int(dead))

        data = pytesseract.image_to_data(dead_img1, output_type=Output.DICT)
        dead = "".join(data["text"])
        dead = dead.replace(",", "")
        if dead.isdigit():
            dead_list.append(int(dead))

        data = pytesseract.image_to_data(dead_img2, output_type=Output.DICT)
        dead = "".join(data["text"])
        dead = dead.replace(",", "")
        if dead.isdigit():
            dead_list.append(int(dead))

        data = pytesseract.image_to_data(dead_img3, output_type=Output.DICT)
        dead = "".join(data["text"])
        dead = dead.replace(",", "")
        if dead.isdigit():
            dead_list.append(int(dead))

        if len(dead_list) > 0:
            dead = mode(dead_list)
        else:
            dead = -1

        print(f"dead: {dead}")
        data_dict["dead"] = dead

        write_to_xlsx(counter, data_dict)
        counter += 1


#############RUN###############
player_count = 300
setup_xlsx()
# takeScreenshot(player_count)
ocr(player_count)